# exporter/excel_writer.py

import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl import Workbook
from datetime import datetime

def write_excel(invoice_df, summary_df=None, skipped_lines=None, mismatched_lines=None, output_dir="outputs"):
    """
    Write the structured Excel file with 2-3 sheets:
    - InvoiceItems
    - BillSummary (optional)
    - Warnings (optional)

    Returns the output path.
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(output_dir, f"invoice_export_{timestamp}.xlsx")

    # Create workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "InvoiceItems"

    # Write InvoiceItems
    for row in dataframe_to_rows(invoice_df, index=False, header=True):
        ws1.append(row)
        print("🧾 Final columns:", df.columns.tolist())

    # Bold headers
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    # Optional: Add summary sheet
    if summary_df is not None:
        ws2 = wb.create_sheet(title="BillSummary")
        for row in dataframe_to_rows(summary_df, index=False, header=True):
            ws2.append(row)
        for cell in ws2[1]:
            cell.font = Font(bold=True)

    # Optional: Add Warnings sheet
    if skipped_lines or mismatched_lines:
        ws3 = wb.create_sheet(title="Warnings")
        ws3.append(["Type", "Line Text", "Details"])
        if skipped_lines:
            for line in skipped_lines:
                ws3.append(["Skipped", line, "Could not parse"])
        if mismatched_lines:
            for line, reason in mismatched_lines:
                ws3.append(["Mismatch", line, reason["Validation"]])
        for cell in ws3[1]:
            cell.font = Font(bold=True)

    wb.save(file_path)
    return file_path
